from io import BytesIO

from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse
from django.core.files.uploadedfile import SimpleUploadedFile

from catalog.models import Category, Priority
from tickets.models import Ticket
from tickets.validators import validate_upload, UploadValidationError


class ReportsExportExcelTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="u1", password="pass")
        cat = Category.objects.create(name="Cat")
        pri = Priority.objects.create(name="Baja")
        Ticket.objects.create(
            title="Test",
            description="d",
            requester=self.user,
            category=cat,
            priority=pri,
        )

    def test_export_excel(self):
        self.client.login(username="u1", password="pass")
        url = reverse("reports_export_excel")
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        from openpyxl import load_workbook

        wb = load_workbook(filename=BytesIO(resp.content))
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        self.assertIn("Código", headers)
        first_row = [cell.value for cell in next(ws.iter_rows(min_row=2, max_row=2))]
        self.assertEqual(first_row[0], "1")


class UploadValidatorTests(TestCase):
    def test_invalid_path_rejected(self):
        class Dummy:
            name = "../../evil.txt"
            size = 1
            content_type = "text/plain"

        with self.assertRaises(UploadValidationError):
            validate_upload(Dummy())

    def test_valid_file(self):
        f = SimpleUploadedFile("ok.txt", b"x", content_type="text/plain")
        try:
            validate_upload(f)
        except UploadValidationError:
            self.fail("validate_upload raised unexpectedly")

